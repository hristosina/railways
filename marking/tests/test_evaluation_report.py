import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

import numpy as np
from openpyxl import load_workbook
from PIL import Image

from evaluation_report import (
    aligned_scenario_view,
    class_id_mapping,
    copy_standard_ultralytics_report,
    discover_scenarios,
    evaluation_progress_percent,
    extract_evaluation,
    publish_standard_ultralytics_report,
    remap_annotation_text,
    resolve_scenario_directory,
    write_report,
)


class EvaluationReportTests(unittest.TestCase):
    def test_class_ids_are_aligned_to_model_order(self):
        dataset_names = ["person", "rails", "train", "vehicle"]
        model_names = ["vehicle", "train", "person", "rails"]
        mapping = class_id_mapping(dataset_names, model_names)
        self.assertEqual(mapping, {0: 2, 1: 3, 2: 1, 3: 0})
        self.assertEqual(
            remap_annotation_text("3 0.5 0.5 1 1\n1 0.2 0.2 0.1 0.1", mapping),
            "0 0.5 0.5 1 1\n3 0.2 0.2 0.1 0.1",
        )

    def test_aligned_view_does_not_modify_source_labels(self):
        with tempfile.TemporaryDirectory() as directory:
            scenario = Path(directory) / "rain"
            (scenario / "images").mkdir(parents=True)
            (scenario / "labels").mkdir()
            (scenario / "images" / "frame.jpg").write_bytes(b"image")
            source_label = scenario / "labels" / "frame.txt"
            source_label.write_text("3 0.5 0.5 1 1", encoding="utf-8")
            with aligned_scenario_view(
                scenario,
                ["person", "rails", "train", "vehicle"],
                ["vehicle", "train", "person", "rails"],
            ) as (view, mapping):
                self.assertEqual(mapping[3], 0)
                self.assertEqual((view / "labels" / "frame.txt").read_text(encoding="utf-8"), "0 0.5 0.5 1 1")
            self.assertEqual(source_label.read_text(encoding="utf-8"), "3 0.5 0.5 1 1")

    def test_progress_changes_inside_scenario(self):
        self.assertEqual(evaluation_progress_percent(0, 3, 0, 100), 0)
        self.assertEqual(evaluation_progress_percent(0, 3, 49, 100), 16)
        self.assertEqual(evaluation_progress_percent(0, 3, 99, 100), 33)
        self.assertEqual(evaluation_progress_percent(1, 3, 49, 100), 50)
        self.assertEqual(evaluation_progress_percent(2, 3, 99, 100), 100)

    def test_scenarios_follow_thesis_order_and_ignore_report(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            for name in ("night", "fog", "fallout", "afternoon", "twilight"):
                (root / name / "images").mkdir(parents=True)
                (root / name / "labels").mkdir()
            (root / "Отчет_тестирования" / "images").mkdir(parents=True)
            (root / "Отчет_тестирования" / "labels").mkdir()
            self.assertEqual(
                [path.name for path in discover_scenarios(root)],
                ["fallout", "fog", "afternoon", "night", "twilight"],
            )

    def test_scenario_folder_accepts_root_or_images(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory) / "rain"
            (root / "images").mkdir(parents=True)
            (root / "labels").mkdir()
            self.assertEqual(resolve_scenario_directory(root), root.resolve())
            self.assertEqual(resolve_scenario_directory(root / "images"), root.resolve())

    def test_standard_report_is_flat_and_prefixed_by_scenario(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "temporary" / "nested"
            source.mkdir(parents=True)
            (source / "BoxPR_curve.png").write_bytes(b"plot")
            copied = copy_standard_ultralytics_report(
                root / "temporary", root / "report", "Ночь/дождь", 2
            )
            self.assertEqual(len(copied), 1)
            self.assertEqual(copied[0].name, "Marking__02_Ночь_дождь__nested__BoxPR_curve.png")
            self.assertEqual(copied[0].read_bytes(), b"plot")
            (root / "report" / "user_file.txt").write_text("keep", encoding="utf-8")
            published = copy_standard_ultralytics_report(
                root / "temporary", root / "staging", "Ночь", 1
            )
            self.assertTrue(published)
            publish_standard_ultralytics_report(root / "staging", root / "report")
            self.assertTrue((root / "report" / "user_file.txt").is_file())
            self.assertFalse(copied[0].exists())

    def test_per_class_metrics_use_ap50_and_ap_arrays(self):
        box = SimpleNamespace(
            ap_class_index=np.array([0, 1, 2, 3]),
            px=np.array([0.0, 0.5, 1.0]),
            prec_values=np.array([[1.0, 0.8, 0.0]] * 4),
            p=np.array([0.1, 0.2, 0.3, 0.4]),
            r=np.array([0.5, 0.6, 0.7, 0.8]),
            ap50=np.array([0.11, 0.22, 0.33, 0.44]),
            ap=np.array([0.01, 0.02, 0.03, 0.04]),
        )
        rows, curves = extract_evaluation(SimpleNamespace(box=box), ["person", "rails", "train", "vehicle"])
        self.assertEqual([row["Класс"] for row in rows], ["Транспорт", "Поезд", "Человек", "Рельсы"])
        self.assertAlmostEqual(rows[0]["mAP@0.5"], 0.44)
        self.assertAlmostEqual(rows[0]["mAP@0.5:0.95"], 0.04)
        self.assertIn("Все классы", curves)

    def test_workbook_contains_thesis_table_and_pr_sheets(self):
        curve = (np.array([0.0, 0.5, 1.0]), np.array([1.0, 0.7, 0.0]))
        with tempfile.TemporaryDirectory() as directory:
            result = {
                "scenario": "Осадки", "images": 10, "precision": 0.8, "recall": 0.7,
                "map50": 0.75, "map5095": 0.5, "inference_ms": 10.0, "fps": 100.0,
                "classes": [{"Класс": "Транспорт", "mAP@0.5": 0.8, "mAP@0.5:0.95": 0.6}],
                "curves": {"Транспорт": curve, "Все классы": curve},
            }
            def make_test_figure(_scenario, _curves, path):
                Image.new("RGB", (20, 20), "white").save(path)

            result["path"] = directory
            with patch("evaluation_report.save_pr_figure", side_effect=make_test_figure):
                path = write_report(
                    Path(directory), "model.pt", 640, "CPU", [result],
                    dataset_class_names=["person", "rails", "train", "vehicle"],
                    model_class_names=["vehicle", "train", "person", "rails"],
                    ultralytics_version="test",
                )
            workbook = load_workbook(path, read_only=False)
            self.assertEqual(workbook["Результаты"]["A2"].value, "Осадки")
            self.assertEqual(workbook["Результаты"]["B3"].value, "Все классы")
            self.assertIn("PR 1 Осадки", workbook.sheetnames)
            self.assertIn("Соответствие классов", workbook.sheetnames)
            self.assertEqual(workbook["Соответствие классов"]["E2"].value, "0 → 2")
            self.assertTrue((Path(directory) / "Marking_PR__01_Осадки.png").is_file())


if __name__ == "__main__":
    unittest.main()

"""Формирование единого отчета о тестировании модели."""

from __future__ import annotations

import re
import shutil
import os
import tempfile
from contextlib import contextmanager
from datetime import datetime
from pathlib import Path
from typing import Iterable

import numpy as np
import pandas as pd
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill


SCENARIO_NAMES = {
    "fallout": "Осадки", "precipitation": "Осадки", "rain": "Осадки", "осадки": "Осадки",
    "fog": "Туман", "туман": "Туман",
    "afternoon": "День", "day": "День", "день": "День",
    "night": "Ночь", "ночь": "Ночь",
    "twilight": "Сумерки", "dusk": "Сумерки", "сумерки": "Сумерки",
}
SCENARIO_ORDER = {name: index for index, name in enumerate(("Осадки", "Туман", "День", "Ночь", "Сумерки"))}
CLASS_NAMES = {
    "vehicle": "Транспорт", "transport": "Транспорт", "транспорт": "Транспорт",
    "train": "Поезд", "поезд": "Поезд",
    "person": "Человек", "human": "Человек", "человек": "Человек",
    "rails": "Рельсы", "rail": "Рельсы", "рельсы": "Рельсы",
}
CLASS_ORDER = {name: index for index, name in enumerate(("Транспорт", "Поезд", "Человек", "Рельсы"))}
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff", ".webp"}


def display_scenario(name: str) -> str:
    return SCENARIO_NAMES.get(name.strip().lower(), name)


def display_class(name: str) -> str:
    return CLASS_NAMES.get(name.strip().lower(), name)


def discover_scenarios(test_root: Path) -> list[Path]:
    """Находит валидные сценарии и располагает их в порядке таблицы диплома."""
    root = Path(test_root).resolve()
    scenarios = []
    for path in (root, *root.rglob("*")):
        if not path.is_dir() or "Отчет_тестирования" in path.parts:
            continue
        if (path / "images").is_dir() and (path / "labels").is_dir():
            scenarios.append(path)

    def sort_key(path: Path):
        shown = display_scenario(path.name)
        return SCENARIO_ORDER.get(shown, len(SCENARIO_ORDER)), shown.casefold()

    return sorted(set(scenarios), key=sort_key)


def resolve_scenario_directory(selected_path: Path) -> Path:
    """Принимает корень сценария или его папку images и проверяет пару images/labels."""
    selected = Path(selected_path).resolve()
    candidate = selected.parent if selected.name.lower() == "images" else selected
    if not (candidate / "images").is_dir() or not (candidate / "labels").is_dir():
        raise ValueError(
            f"В папке сценария «{selected}» должны находиться подпапки images и labels."
        )
    return candidate


def normalize_class_names(names) -> list[str]:
    """Нормализует список или словарь имен классов с сохранением порядка ID."""
    if isinstance(names, dict):
        names = [names[key] for key in sorted(names, key=lambda value: int(value))]
    if not isinstance(names, (list, tuple)) or not all(isinstance(name, str) for name in names):
        raise ValueError("Некорректный список классов.")
    return [name.strip() for name in names]


def class_id_mapping(dataset_names, model_names) -> dict[int, int]:
    """Возвращает соответствие ID датасета → ID модели по именам классов."""
    source = normalize_class_names(dataset_names)
    target = normalize_class_names(model_names)
    source_keys = [name.casefold() for name in source]
    target_keys = [name.casefold() for name in target]
    if len(set(source_keys)) != len(source_keys) or len(set(target_keys)) != len(target_keys):
        raise ValueError("Названия классов должны быть уникальными.")
    if set(source_keys) != set(target_keys):
        only_dataset = [source[i] for i, key in enumerate(source_keys) if key not in set(target_keys)]
        only_model = [target[i] for i, key in enumerate(target_keys) if key not in set(source_keys)]
        raise ValueError(
            "Классы датасета и модели не совпадают. "
            f"Только в data.yaml: {only_dataset or '—'}; только в модели: {only_model or '—'}."
        )
    target_ids = {key: index for index, key in enumerate(target_keys)}
    return {source_id: target_ids[key] for source_id, key in enumerate(source_keys)}


def remap_annotation_text(text: str, mapping: dict[int, int], source: Path | None = None) -> str:
    """Перенумеровывает первый столбец YOLO-разметки."""
    output = []
    for line_number, line in enumerate(text.splitlines(), start=1):
        stripped = line.strip()
        if not stripped:
            continue
        parts = stripped.split(maxsplit=1)
        try:
            class_id = int(parts[0])
        except ValueError as exc:
            raise ValueError(f"Некорректный ID класса в {source or 'разметке'}, строка {line_number}.") from exc
        if class_id not in mapping:
            raise ValueError(f"Неизвестный ID класса {class_id} в {source or 'разметке'}, строка {line_number}.")
        suffix = f" {parts[1]}" if len(parts) == 2 else ""
        output.append(f"{mapping[class_id]}{suffix}")
    return "\n".join(output)


@contextmanager
def aligned_scenario_view(scenario_dir: Path, dataset_names, model_names):
    """Создает временное представление с ID классов в порядке обученной модели."""
    scenario_dir = resolve_scenario_directory(scenario_dir)
    mapping = class_id_mapping(dataset_names, model_names)
    if all(source_id == target_id for source_id, target_id in mapping.items()):
        yield scenario_dir, mapping
        return

    with tempfile.TemporaryDirectory(prefix=".marking_class_alignment_", dir=scenario_dir.parent) as directory:
        view = Path(directory) / "scenario"
        images_out, labels_out = view / "images", view / "labels"
        images_out.mkdir(parents=True)
        labels_out.mkdir()
        for image in (scenario_dir / "images").iterdir():
            if not image.is_file() or image.suffix.lower() not in IMAGE_EXTENSIONS:
                continue
            destination = images_out / image.name
            try:
                os.link(image, destination)
            except OSError:
                shutil.copy2(image, destination)
            label = scenario_dir / "labels" / f"{image.stem}.txt"
            if label.is_file():
                remapped = remap_annotation_text(
                    label.read_text(encoding="utf-8"), mapping, source=label
                )
                (labels_out / label.name).write_text(remapped, encoding="utf-8")
        yield view, mapping


def ordered_classes(class_names: Iterable[str]) -> list[tuple[int, str, str]]:
    result = [(index, str(name), display_class(str(name))) for index, name in enumerate(class_names)]
    return sorted(result, key=lambda item: (CLASS_ORDER.get(item[2], len(CLASS_ORDER)), item[0]))


def extract_evaluation(metrics, class_names: list[str]) -> tuple[list[dict], dict[str, tuple[np.ndarray, np.ndarray]]]:
    """Извлекает правильно сопоставленные метрики классов и PR-кривые."""
    box = metrics.box
    class_positions = {int(class_id): pos for pos, class_id in enumerate(np.asarray(box.ap_class_index).tolist())}
    recall_axis = np.asarray(box.px, dtype=float)
    precision_values = np.asarray(box.prec_values, dtype=float)
    rows, curves, available_curves = [], {}, []
    for class_id, source_name, shown_name in ordered_classes(class_names):
        pos = class_positions.get(class_id)
        if pos is None:
            precision = recall = map50 = map5095 = 0.0
            curve = np.full_like(recall_axis, np.nan, dtype=float)
        else:
            precision = float(box.p[pos])
            recall = float(box.r[pos])
            map50 = float(box.ap50[pos])
            map5095 = float(box.ap[pos])
            curve = np.asarray(precision_values[pos], dtype=float)
            available_curves.append(curve)
        rows.append({
            "class_id": class_id, "Класс": shown_name, "Исходное имя класса": source_name,
            "mAP@0.5": map50, "mAP@0.5:0.95": map5095,
            "Precision": precision, "Recall": recall,
        })
        curves[shown_name] = (recall_axis, curve)
    mean_curve = np.nanmean(np.vstack(available_curves), axis=0) if available_curves else np.zeros_like(recall_axis)
    curves["Все классы"] = (recall_axis, mean_curve)
    return rows, curves


def count_images(images_dir: Path) -> int:
    return sum(1 for path in images_dir.iterdir() if path.is_file() and path.suffix.lower() in IMAGE_EXTENSIONS)


def evaluation_progress_percent(scenario_index: int, total_scenarios: int,
                                batch_index: int, total_batches: int) -> int:
    """Общий прогресс с учетом текущего пакета внутри сценария (индексы с нуля)."""
    if total_scenarios <= 0 or total_batches <= 0:
        return 0
    completed_scenarios = max(0, min(scenario_index, total_scenarios))
    completed_batches = max(0, min(batch_index + 1, total_batches))
    fraction = (completed_scenarios + completed_batches / total_batches) / total_scenarios
    return max(0, min(100, int(fraction * 100)))


def safe_filename(value: str) -> str:
    return re.sub(r'[<>:"/\\|?*]+', "_", value).strip(" .") or "Сценарий"


def copy_standard_ultralytics_report(source_dir: Path, output_dir: Path,
                                     scenario_name: str, scenario_index: int) -> list[Path]:
    """Копирует стандартный отчет в одну папку и добавляет сценарий к именам."""
    source_dir, output_dir = Path(source_dir), Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    prefix = f"Marking__{scenario_index:02d}_{safe_filename(scenario_name)[:60]}__"
    copied = []
    if not source_dir.is_dir():
        return copied
    for source in source_dir.rglob("*"):
        if not source.is_file():
            continue
        relative_name = "__".join(source.relative_to(source_dir).parts)
        destination = output_dir / f"{prefix}{relative_name}"
        shutil.copy2(source, destination)
        copied.append(destination)
    return copied


def publish_standard_ultralytics_report(staging_dir: Path, output_dir: Path) -> list[Path]:
    """Атомарно по смыслу заменяет только ранее созданные приложением файлы."""
    staging_dir, output_dir = Path(staging_dir), Path(output_dir)
    new_files = [path for path in staging_dir.glob("Marking__*") if path.is_file()]
    if not new_files:
        return []
    output_dir.mkdir(parents=True, exist_ok=True)
    for old_file in output_dir.glob("Marking__*"):
        if old_file.is_file():
            old_file.unlink()
    published = []
    for source in new_files:
        destination = output_dir / source.name
        shutil.copy2(source, destination)
        published.append(destination)
    return published


def safe_sheet_title(value: str) -> str:
    return re.sub(r'[\[\]:*?/\\]+', "_", value)[:31] or "Лист"


def save_pr_figure(scenario: str, curves: dict[str, tuple[np.ndarray, np.ndarray]], path: Path) -> None:
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    fig, ax = plt.subplots(figsize=(9, 6), dpi=150)
    for class_name, (recall, precision) in curves.items():
        if class_name == "Все классы":
            ax.plot(recall, precision, linewidth=3, linestyle="--", color="#1677ff", label=class_name)
        elif not np.isnan(precision).all():
            ax.plot(recall, precision, linewidth=1.8, label=class_name)
    ax.set_title(f"Precision–Recall: {scenario}", fontweight="bold")
    ax.set_xlabel("Полнота (Recall)")
    ax.set_ylabel("Точность (Precision)")
    ax.set_xlim(0, 1)
    ax.set_ylim(0, 1.02)
    ax.grid(True, alpha=0.22)
    ax.legend(loc="lower left")
    fig.tight_layout()
    fig.savefig(path, bbox_inches="tight")
    plt.close(fig)


def write_report(report_dir: Path, model_path: str, imgsz: int, device_name: str,
                 scenario_results: list[dict], dataset_class_names=None,
                 model_class_names=None, ultralytics_version="") -> Path:
    """Создает одну книгу Excel и по одному PR-графику на сценарий."""
    report_dir = Path(report_dir)
    report_dir.mkdir(parents=True, exist_ok=True)
    workbook_path = report_dir / "Отчет_тестирования.xlsx"
    result_rows, summary_rows, pr_rows, figure_paths = [], [], [], {}
    for old_figure in report_dir.glob("Marking_PR__*.png"):
        if old_figure.is_file():
            old_figure.unlink()
    for scenario_index, result in enumerate(scenario_results, start=1):
        scenario = result["scenario"]
        for row in result["classes"]:
            result_rows.append({"Сценарий": scenario, "Класс": row["Класс"],
                                "mAP@0.5": row["mAP@0.5"], "mAP@0.5:0.95": row["mAP@0.5:0.95"]})
        result_rows.append({"Сценарий": scenario, "Класс": "Все классы",
                            "mAP@0.5": result["map50"], "mAP@0.5:0.95": result["map5095"]})
        summary_rows.append({
            "Сценарий": scenario, "Изображений": result["images"], "Precision": result["precision"],
            "Recall": result["recall"], "mAP@0.5": result["map50"], "mAP@0.5:0.95": result["map5095"],
            "Инференс, мс/изобр.": result["inference_ms"], "FPS": result["fps"],
        })
        for class_name, (recall, precision) in result["curves"].items():
            for recall_value, precision_value in zip(recall, precision):
                pr_rows.append({"Сценарий": scenario, "Класс": class_name, "Recall": float(recall_value),
                                "Precision": None if np.isnan(precision_value) else float(precision_value)})
        figure_path = report_dir / f"Marking_PR__{scenario_index:02d}_{safe_filename(scenario)[:80]}.png"
        save_pr_figure(scenario, result["curves"], figure_path)
        figure_paths[scenario] = figure_path

    parameters = [
        ("Дата формирования", datetime.now().strftime("%d.%m.%Y %H:%M")),
        ("Модель", str(Path(model_path).resolve())),
        ("Размер изображения", imgsz), ("Устройство", device_name), ("Сценариев", len(scenario_results)),
        ("Порог confidence", 0.001), ("Порог NMS IoU", 0.5),
        ("Метрики AP", "IoU 0.5 и среднее IoU 0.5:0.95"),
    ]
    if ultralytics_version:
        parameters.append(("Версия Ultralytics", ultralytics_version))
    if dataset_class_names is not None:
        parameters.append(("Порядок классов data.yaml", ", ".join(normalize_class_names(dataset_class_names))))
    if model_class_names is not None:
        parameters.append(("Порядок классов модели", ", ".join(normalize_class_names(model_class_names))))
    parameters.extend(
        (f"Папка сценария «{result['scenario']}»", result["path"])
        for result in scenario_results
    )
    alignment_rows = []
    if dataset_class_names is not None and model_class_names is not None:
        source_names = normalize_class_names(dataset_class_names)
        target_names = normalize_class_names(model_class_names)
        mapping = class_id_mapping(source_names, target_names)
        alignment_rows = [
            {
                "ID в data.yaml": source_id,
                "Класс в data.yaml": source_names[source_id],
                "ID модели": target_id,
                "Класс модели": target_names[target_id],
                "Перенумерация": "Не требуется" if source_id == target_id else f"{source_id} → {target_id}",
            }
            for source_id, target_id in mapping.items()
        ]
        parameters.append((
            "Перенумерация ID классов",
            "Выполнена во временной копии разметки"
            if any(source_id != target_id for source_id, target_id in mapping.items())
            else "Не требуется",
        ))
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        pd.DataFrame(result_rows).to_excel(writer, sheet_name="Результаты", index=False)
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Сводка", index=False)
        pd.DataFrame(parameters, columns=["Параметр", "Значение"]).to_excel(writer, sheet_name="Параметры", index=False)
        if alignment_rows:
            pd.DataFrame(alignment_rows).to_excel(writer, sheet_name="Соответствие классов", index=False)
        pd.DataFrame(pr_rows).to_excel(writer, sheet_name="PR данные", index=False)
        workbook = writer.book
        header_fill = PatternFill("solid", fgColor="2F80ED")
        alternate_fill = PatternFill("solid", fgColor="EAF3FF")
        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            sheet.sheet_view.showGridLines = False
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for column in sheet.columns:
                letter = column[0].column_letter
                sheet.column_dimensions[letter].width = min(55, max(12, max(len(str(cell.value or "")) for cell in column) + 2))

        results_sheet = workbook["Результаты"]
        results_sheet.freeze_panes = "C2"
        results_sheet.column_dimensions["A"].width = 18
        results_sheet.column_dimensions["B"].width = 22
        for row_index in range(2, results_sheet.max_row + 1):
            if results_sheet.cell(row_index, 2).value == "Все классы":
                for cell in results_sheet[row_index]:
                    cell.fill = alternate_fill
                    cell.font = Font(bold=True)
            for column_index in (3, 4):
                results_sheet.cell(row_index, column_index).number_format = "0.000"
        results_sheet.conditional_formatting.add(
            f"C2:D{results_sheet.max_row}",
            ColorScaleRule(start_type="num", start_value=0, start_color="F8696B",
                           mid_type="num", mid_value=0.5, mid_color="FFEB84",
                           end_type="num", end_value=1, end_color="63BE7B"),
        )
        for row_index in range(2, workbook["Сводка"].max_row + 1):
            for column_index in range(3, 9):
                workbook["Сводка"].cell(row_index, column_index).number_format = "0.000"

        for index, result in enumerate(scenario_results, start=1):
            scenario = result["scenario"]
            sheet = workbook.create_sheet(safe_sheet_title(f"PR {index} {scenario}"))
            sheet.sheet_view.showGridLines = False
            sheet["A1"] = f"Precision–Recall — {scenario}"
            sheet["A1"].font = Font(size=16, bold=True, color="1F3555")
            sheet["A3"] = "График сформирован автоматически по результатам одного тестового прогона."
            sheet["A3"].alignment = Alignment(wrap_text=True)
            sheet.column_dimensions["A"].width = 70
            image = ExcelImage(str(figure_paths[scenario]))
            image.width, image.height = 900, 600
            sheet.add_image(image, "A5")
    return workbook_path
